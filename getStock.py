programVersion = "version 1.0"
#author https://github.com/papatekken/
#create date 2020-07-10


from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date

import yfinance as yf
from datetime import datetime, timedelta

print (str(datetime.now())+' Start')



def captureData(wsdata,value,row):
	print(value+":START")
	
	#init value
	currPrice=0
	openPrice=0
	closePrice=0
	highPrice=0
	lowPrice=0
	high52Price=0
	low52Price=0
	day50avg=0
	day150avg=0
	day200avg=0
	day200avg2=0
	day200avg3=0
	day200avg4=0
	day200avg5=0
	
	#get online
	try:
		ticker = yf.Ticker(value)
		data = yf.download(value, start=today, end=historytoday)
	
		currPrice=data["Close"].tail(1).min()
		openPrice=data["Open"].tail(1).min()
		closePrice=data["Close"].tail(1).min()
		highPrice=data["High"].tail(1).min()
		lowPrice=data["Low"].tail(1).min()
		high52Price=ticker.history(period="ytd")["High"].max()
		low52Price=ticker.history(period="ytd")["Low"].min()	
		day50avg=ticker.history(start=day1200, end=valtoday)["Close"].tail(50).mean()
		day150avg=ticker.history(start=day1200, end=valtoday)["Close"].tail(150).mean()
		day200avg=ticker.history(start=day1200, end=valtoday)["Close"].tail(200).mean()
		day200avg2=ticker.history(start=day1200, end=valtoday)["Close"].tail(400).head(200).mean()
		day200avg3=ticker.history(start=day1200, end=valtoday)["Close"].tail(600).head(200).mean()
		day200avg4=ticker.history(start=day1200, end=valtoday)["Close"].tail(800).head(200).mean()
		day200avg5=ticker.history(start=day1200, end=valtoday)["Close"].tail(1000).head(200).mean()		
	except:
		pass	

	#update to excel
	wsdata.cell(row,2).value =currPrice
	wsdata.cell(row,3).value =openPrice
	wsdata.cell(row,4).value =closePrice
	wsdata.cell(row,5).value =highPrice
	wsdata.cell(row,6).value =lowPrice
	wsdata.cell(row,7).value =high52Price
	wsdata.cell(row,8).value =low52Price
	wsdata.cell(row,9).value =day50avg
	wsdata.cell(row,10).value =day150avg
	wsdata.cell(row,11).value =day200avg
	wsdata.cell(row,12).value =day200avg2
	wsdata.cell(row,13).value =day200avg3
	wsdata.cell(row,14).value =day200avg4
	wsdata.cell(row,15).value =day200avg5
	
	if(currPrice>day150avg and currPrice > day200avg):
		wsdata.cell(row,16).value = "Y"
	else:	
		wsdata.cell(row,16).value = ""
		
	if(day150avg>day200avg):
		wsdata.cell(row,17).value = "Y"
	else:	
		wsdata.cell(row,17).value = ""		
		
	if(day200avg>day200avg2 and day200avg2>day200avg3):
		wsdata.cell(row,18).value = "Y"
	else:	
		wsdata.cell(row,18).value = ""				
	if(day50avg>day150avg and day150avg > day200avg):
		wsdata.cell(row,19).value = "Y"
	else:	
		wsdata.cell(row,19).value = ""		
	if(currPrice>day50avg):
		wsdata.cell(row,20).value = "Y"
	else:	
		wsdata.cell(row,20).value = ""			
		
	if((currPrice/low52Price)>1.3):
		wsdata.cell(row,21).value = "Y"
	else:	
		wsdata.cell(row,21).value = ""
		
	if((currPrice/high52Price)>0.75):
		wsdata.cell(row,22).value = "Y"
	else:	
		wsdata.cell(row,22).value = ""
	wsdata.cell(row,23).value = '=countif(P'+str(row)+':V'+str(row)+',"Y")'
	print(value+":Done")
	
print("*** Capture stock info online ***")
print(programVersion)

##assign excel worksbook and worksheet
wb = load_workbook('data.xlsx')
wsdata =wb['list']

##current day
historytoday =date.today().strftime('%Y-%m-%d')


##last trade day
valtoday =  date.today()
while(valtoday.weekday()==5 or valtoday.weekday()==6 or (valtoday.weekday()==1 and datetime.utcnow().hour<=8)):
	valtoday =valtoday-timedelta(days=1)
today=valtoday.strftime('%Y-%m-%d')


day1200=(valtoday - timedelta(days=1200)).strftime('%Y-%m-%d')
number_rows = wsdata.max_row+1


##loop each stock ticker symbol and call function captureData to get information online
for row in range(2,number_rows):
		value = wsdata.cell(row, 1).value
		if value != None:
			captureData(wsdata,value,row)		
		else:
			break

##save and end program			
wb.save('data'+historytoday+'.xlsx')	
print (str(datetime.now())+' Finished')
print("Finished")

